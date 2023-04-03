import openpyxl

class trainee:
    def __init__(self, id, name, course, degree, workExperience):
        self.id = id
        self.name = name
        self.course = course
        self.degree = degree
        self.workExperience = workExperience

    def addTrainee(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        last_row = ws.max_row
        ws.cell(row=last_row+1, column=1, value=self.id)
        ws.cell(row=last_row+1, column=2, value=self.name)
        ws.cell(row=last_row+1, column=3, value=self.course)
        ws.cell(row=last_row+1, column=4, value=self.degree)
        ws.cell(row=last_row+1, column=5, value=self.workExperience)
        wb.save('MasterRecord.xlsx')

    def deleteTrainee(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value == self.id:
                    ws.delete_rows(cell.row, 1)
        wb.save('MasterRecord.xlsx')

    def updateTrainee(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        for row in ws.iter_rows(min_row=2, max_col=5):
            if row[0].value == self.id:
                row[1].value = self.name
                row[2].value = self.course
                row[3].value = self.degree
                row[4].value = self.workExperience
                break
        else:
            print("Trainee not found.")
        wb.save('MasterRecord.xlsx')

def traineeMenu():

    while True:
        print("1. Create trainee")
        print("2. Delete trainee")
        print("3. View trainee")
        print("4. View all trainees")
        print("5. Update trainee")
        print("0. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            id = input("Enter trainee ID: ")
            name = input("Enter trainee name: ")
            course = input("Enter course: ")
            degree = input("Enter degree: ")
            work_exp = input("Enter work experience: ")
            t = trainee(id, name, course, degree, work_exp)
            t.addTrainee()

        elif choice == '2':
            id = input("Enter trainee ID: ")
            t = trainee(id, '', '', '', '')
            t.deleteTrainee()

        elif choice == '3':
            id = input("Enter trainee ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    print(f"ID: {row[0].value}")
                    print(f"Name: {row[1].value}")
                    print(f"Course: {row[2].value}")
                    print(f"Degree: {row[3].value}")
                    print(f"Work Experience: {row[4].value}")
                    break
            else:
                print("Trainee not found.")

        elif choice == '4':
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(2, ws.max_row):
                if all(c.value is None for c in row):
                    break
                else:
                    print("ID: ",row[0].value)
                    print("Name: ",row[1].value)
                    print("Course: ",row[2].value)
                    print("Degree: ",row[3].value)
                    print("Work Experience: ",row[4].value)
                    print()

        elif choice == '5':
            id = input("Enter trainee ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    name = input("Enter new name for trainee with ID: ")
                    course = input("Enter new course for trainee with ID: ")
                    degree = input("Enter new degree for trainee with ID: ")
                    work_exp = input("Enter new work experience for trainee with ID: ")
                    t = trainee(id, name, course, degree, work_exp)
                    t.updateTrainee()
                    break
            else:
                print("Trainee not found.")

        elif choice == '0':
            break
        else:
            print("Invalid choice. Please try again.")