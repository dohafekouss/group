import openpyxl

class Manager():
    def __init__(self, id, fullName, email, phoneNumber):
        self.id=id
        self.fullName=fullName
        self.email=email
        self.phoneNumber=phoneNumber

    def addManager(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self.id)
            ws.cell(row=last_row + 1, column=2, value=self.fullName)
            ws.cell(row=last_row + 1, column=3, value=self.email)
            ws.cell(row=last_row + 1, column=4, value=self.phoneNumber)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving manager to Excel file: {e}")

    def updateManager(self, fullName=None, email=None, phoneNumber=None):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            for row in ws.iter_rows(min_row=2, max_col=6):
                if row[0].value == self._trainer_id:
                    if fullName is not None:
                        row[1].value = fullName
                    if email is not None:
                        row[2].value = email
                    if phoneNumber is not None:
                        row[3].value = phoneNumber
                    break
            else:
                print("Manager not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating manager in Excel file: {e}")

    def deleteManager(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            for row in ws.iter_rows(min_row=2, max_col=6):
                if row[0].value == self.id:
                    ws.delete_rows(row[0].row, 1)
                    break
            else:
                print("Manager not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting manager from Excel file: {e}")

def managerMenu():

    while True:
        print("1. Create manager")
        print("2. Delete manager")
        print("3. View manager")
        print("4. View all managers")
        print("5. Update manager")
        print("6. Exit")

        choice = input("Enter your choice: ")
        print()

        if choice == '1':
            id = input("Enter manager ID: ")
            fullName = input("Enter manager description: ")
            email = input("Enter manager email address: ")
            phoneNumber = input("Enter manager phone number: ")
            t = Manager(id, fullName, email, phoneNumber)
            t.addManager()

        elif choice == '2':
            id = input("Enter manager ID: ")
            t = Manager(id, '', '', '')
            t.deleteManager()

        elif choice == '3':
            id = input("Enter course ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['MangerDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    print(f"ID: ",row[0].value)
                    print(f"Full Name: ",row[1].value)
                    print(f"Email Address: ",row[2].value)
                    print(f"Phone Number: ",row[3].value)
                    break
            else:
                print("Manager doesn't exist!")

        elif choice == '4':
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ManagerDetails']
            for row in ws.iter_rows(2, ws.max_row):
                if all(c.value is None for c in row):
                    break
                else:
                    print("ID: ",row[0].value)
                    print("Full Name: ",row[1].value)
                    print("Email Address: ",row[2].value)
                    print("Phone Number: ",row[3].value)
                    print()

        elif choice == '5':
            id = input("Enter manager ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ManagerDetails']
            for row in ws.iter_rows(1, ws.max_row):
                if row[0].value == id:
                    fullName = input("Enter new name: ")
                    email = input("Enter new email address: ")
                    phoneNumber = input("Enter new phone number: ")
                    t = Manager(id, fullName, email, phoneNumber)
                    t.updateManager()
                    break
            else:
                print("Manager doesn't exist!")

        elif choice == '6':
            break
        else:
            print("Invalid choice! Try again.")