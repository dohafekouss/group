import openpyxl

def mappingCourseTrainerMenu():
    choice=11
    while (choice!=0):
        print("1. Add trainer to a course")
        print("2. View sheet")
        print("3. Exit")

        choice=input("Enter your choice: ")

        if (choice=='1'):
            trainerID = input("Please enter trainer's id: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['TrainerDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, Trainer doesn't exist")
                    continue

            courseID = input("Please enter course's id: ")
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, Course doesn't exist")
                    continue

            ws = wb["MappingCourseTrainer"]
            new_row = (courseID, trainerID)
            ws.append(new_row)
            wb.save("MasterRecord.xlsx")
        elif (choice=='2'):
            break
        elif (choice=='3'):
            break
