import openpyxl

def mappingCourseTraineeMenu():
    choice=11
    while (choice!=0):
        print("1. Add trainee to a course")
        print("2. View sheet")
        print("3. Exit")

        choice=input("Enter your choice: ")

        if (choice=='1'):
            traineeID = input("Please enter trainee's id: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, Trainee doesn't exist")
                    continue

            courseID = input("Please enter course's id: ")
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, Course doesn't exist")
                    continue

            ws = wb["MappingCourseTrainee"]
            new_row = (courseID, traineeID)
            ws.append(new_row)
            wb.save("MasterRecord.xlsx")
        elif (choice=='2'):
            break
        elif (choice=='3'):
            break
