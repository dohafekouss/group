import openpyxl
class course:
    def __init__(self, id, description):
        self.id = id
        self.description=description

    def addCourse(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        last_row = ws.max_row
        ws.cell(row=last_row+1, column=1, value=self.id)
        ws.cell(row=last_row+1, column=2, value=self.description)
        wb.save('MasterRecord.xlsx')

    def deleteCourse(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        for row in ws.iter_rows(min_row=2, max_col=5):
            for cell in row:
                if cell.value == self.id:
                    ws.delete_rows(cell.row, 1)
        wb.save('MasterRecord.xlsx')

    def updateCourse(self):
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        for row in ws.iter_rows(min_row=2, max_col=5):
            if row[0].value == self.id:
                row[1].value = self.description
                break
        else:
            print("Course doesn't exist.")
        wb.save('MasterRecord.xlsx')


def courseMenu():

    while True:
        print("1. Create course")
        print("2. Delete course")
        print("3. View course")
        print("4. View all courses")
        print("5. Update course")
        print("6. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            id = input("Enter course ID: ")
            description = input("Enter course description: ")
            t = course(id, description)
            t.addCourse()

        elif choice == '2':
            id = input("Enter course ID: ")
            t = course(id, '')
            t.deleteCourse()

        elif choice == '3':
            id = input("Enter course ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    print(f"ID: {row[0].value}")
                    print(f"Description: {row[1].value}")
                    break
            else:
                print("Trainee not found.")

        elif choice == '4':
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['CourseDetails']
            for row in ws.iter_rows(2, ws.max_row):
                if all(c.value is None for c in row):
                    break
                else:
                    print("ID: ",row[0].value)
                    print("Description: ",row[1].value)
                    print()
                    #print()

        elif choice == '5':
            id = input("Enter course ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    description = input("Enter new description for course: ")
                    t = course(id, description)
                    t.updateCourse()
                    break
            else:
                print("Course doesn't exist")

        elif choice == '6':
            break
        else:
            print("Invalid choice! Try again.")