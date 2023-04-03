import openpyxl

class Trainer:
    def __init__(self, trainer_id, full_name, email, phone_number):
        self.id = trainer_id
        self.fullName = full_name
        self.email = email
        self.phoneNumber = phone_number

    def addTrainer(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainers']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self.id)
            ws.cell(row=last_row + 1, column=2, value=self.fullName)
            ws.cell(row=last_row + 1, column=3, value=self.email)
            ws.cell(row=last_row + 1, column=4, value=self.phoneNumber)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving trainer to Excel file: {e}")

    def deleteTrainer(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['TrainerDetails']
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == self.id:
                        ws.delete_rows(cell.row, 1)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error detected! Trainer has NOT been deleted: {e}")

    def updateTrainer(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['TrainerDetails']
            for row in ws.iter_rows(min_row=2, max_col=4):
                if row[0].value == self.id:
                    row[1].value = self.fullName
                    row[2].value = self.email
                    row[3].value = self.phoneNumber
                    break
            else:
                print("Trainer not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error detected! Trainer has NOT been updated: {e}")


def trainerMenu():

    while True:
        print("1. Create trainer")
        print("2. Delete trainer")
        print("3. View trainer")
        print("4. View all trainers")
        print("5. Update trainer")
        print("6. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            id = input("Enter trainer ID: ")
            fullName = input("Enter trainer full name: ")
            email = input("Enter trainer email address: ")
            phoneNumber = input("Enter phone number: ")
            t = Trainer(id, fullName, email, phoneNumber)
            t.addTrainer()

        elif choice == '2':
            id = input("Enter trainer ID: ")
            t = Trainer(id, '', '', '')
            t.deleteTrainer()

        elif choice == '3':
            id = input("Enter trainer ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['TrainerDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    print(f"ID: {row[0].value}")
                    print(f"Full Name: {row[1].value}")
                    print(f"Email Address: {row[2].value}")
                    print(f"Phone Number: {row[3].value}")
                    break
            else:
                print("Trainer not found.")

        elif choice == '4':
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['TrainerDetails']
            for row in ws.iter_rows(2, ws.max_row):
                if all(c.value is None for c in row):
                    break
                else:
                    print(f"ID: ",row[0].value)
                    print(f"Full Name: ",row[1].value)
                    print(f"Email Address: ",row[2].value)
                    print(f"Phone Number: ",row[3].value)
                    print()

        elif choice == '5':
            id = input("Enter trainer ID: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == id:
                    fullName = input("Enter new name for trainer: ")
                    emailAddress = input("Enter new email address for trainer: ")
                    phoneNumber = input("Enter new phone number for trainer: ")
                    t = Trainer(id, fullName, emailAddress, phoneNumber)
                    t.updateTrainer()
                    break
            else:
                print("Trainer doesn't exist!")
        elif choice == '6':
            break
        else:
            print("Invalid choice. Please try again.")