import openpyxl

def sessionMenu():
    choice = 11

    while (choice != '0'):
        print("Session Menu: ")
        print("1: Add Session")
        print("2: Add Absentee")
        print("3: Exit")

        choice = input("Please enter your choice: ")
        if (choice=='1'):
            sessionDate = input("Please enter the session date: ")
            classID = input("Please enter the session id: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, class doesn't exist")
                    continue
            sheets = wb.sheetnames
            if (classID + " " + sessionDate) in sheets:
                print("Error, class session already created for that date")
                continue

            wb.create_sheet(classID + " " + sessionDate)
            wb.save("MasterRecord.xlsx")

            startTime = input("Please enter the start time of the session: ")
            endTime = input("Please enter the end time of the session: ")

            ws = wb[classID + " " + sessionDate]
            new_row = (sessionDate,startTime, endTime,classID)
            ws.append(new_row)
            wb.save("MasterRecord.xlsx")

            wr = wb["MappingCourseTrainee"]
            for i in range(1, wr.max_row+1):
                if(wr.cell(row=i,column=1).value==classID):
                    x = wr.cell(row=i,column=2).value
            wr = wb["ListOfTrainees"]
            for i in range(1, wr.max_row+1):
                if (wr.cell(row=i, column=1).value == x):
                    new_row = (x, wr.cell(row=i, column=2).value, "P")
                    ws.append(new_row)
                    wb.save("MasterRecord.xlsx")


        elif (choice=='2'):
            sessionDate = input("Please enter the date of the session: ")
            classID = input("Please enter the id of the class of the session: ")
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['CourseDetails']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if (row[0].value == -1):
                    print("Error, class doesn't exist")
                    continue

            sheets = wb.sheetnames
            if not ((classID + " " + sessionDate) in sheets):
                print("Error, Class sesson not created for that date")
                continue

            inputVal = input("Please input the list of student IDs: ")
            studentIDList = inputVal.split(" ")

            ws = wb[classID + " " + sessionDate]
            students = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]

            for x in range(0, len(students)):
                if (students[x] in studentIDList):
                    ws.cell(row=2 + x, column=2, value="Absent")
                    wb.save("MasterRecord.xlsx")

        elif (choice=='3'):
            break


